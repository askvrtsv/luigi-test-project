import abc
import csv
import pathlib
import typing as t

from openpyxl import load_workbook

from newproducts.models import (
    ProcessedProduct,
    Product,
    ProductStatus,
    RepoProduct,
    ScrapedProduct,
    StoreConfig,
)
from newproducts.settings import SCRAPED_DIR


class DataSource(abc.ABC):
    filename: pathlib.Path

    @abc.abstractmethod
    def read_items(self) -> t.Iterable[dict]:
        raise NotImplementedError


class CsvSource(DataSource):
    filename: pathlib.Path
    dialect = "excel"

    def read_items(self) -> t.Iterable[dict]:
        with self.filename.open("r", encoding="utf-8") as f:
            yield from csv.DictReader(f, dialect=self.dialect)


class ExcelSource(DataSource):
    def read_items(self) -> t.Iterable[dict]:
        workbook = load_workbook(self.filename, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return None
        for values in rows:
            product = dict(zip(header, values))
            yield product


class ProductReader(abc.ABC):
    def __init__(self, filename: pathlib.Path) -> None:
        self.filename = filename

    @abc.abstractmethod
    def get_products(self) -> t.Iterable[Product]:
        raise NotImplementedError

    def __iter__(self):
        yield from self.get_products()


class RepoProductReader(ProductReader, ExcelSource):
    def get_products(self) -> t.Iterable[RepoProduct]:
        for product in self.read_items():
            yield RepoProduct(
                store_name=product["Магазин"],
                sku=product["Артикул"],
                title=product["Наименование"],
                image=product["Картинка"],
                category=product["Категория"],
                brand=product["Бренд"],
                label=product["Новые"],
            )


class UniqueProductReader(ProductReader, CsvSource):
    def get_products(self) -> t.Iterable[Product]:
        for product in self.read_items():
            yield Product(
                store_name=product["store"],
                sku=product["sku"],
            )


class ScrapedProductReader(ProductReader, CsvSource):
    dialect = "excel-tab"

    def __init__(self, filename: pathlib.Path, store: StoreConfig) -> None:
        super().__init__(filename)
        self.store = store

    def get_products(self) -> t.Iterable[ScrapedProduct]:
        for product in self.read_items():
            values = list(product.values())
            yield ScrapedProduct(
                store_name=self.store.name,
                sku=values[self.store.sku_col - 1],
                title=values[self.store.title_col - 1],
                image=(
                    values[self.store.image_col - 1]
                    if self.store.image_col is not None
                    else None
                ),
                category=(
                    values[self.store.category_col - 1]
                    if self.store.category_col is not None
                    else None
                ),
                brand=(
                    values[self.store.brand_col - 1]
                    if self.store.brand_col is not None
                    else None
                ),
                label="",
            )


class ProcessedProductReader(ProductReader, CsvSource):
    def __init__(self, filename: pathlib.Path, predicate: t.Callable) -> None:
        self.filename = filename
        self.predicate = predicate

    def get_products(self) -> t.Iterable[ProcessedProduct]:
        for data in self.read_items():
            product = ProcessedProduct(
                store_name=data["store"],
                sku=data["sku"],
                title=data["title"],
                image=data["image"],
                category=data["category"],
                brand=data["brand"],
                label=data["label"],
                status=ProductStatus(data["status"]),
            )
            if self.predicate(product):
                yield product
